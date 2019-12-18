======== Python 2.7 =================
import xlrd
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

from bs4 import BeautifulSoup as soup
import re as regex  # python regex module
import time
import xlwt
import pandas
from openpyxl import load_workbook

# System.setProperty("webdriver.chrome.driver", "D:/apps/WinPython-64bit-3.5.3.1Qt5/bs4_tuts/operadriver_win32/operadriver_win32/operadriver.exe");
# WebDriver driver = new ChromeDriver();
#cdriver.get("https://www.google.com");
xls_file1 = xlrd.open_workbook("E:/css/crawling/code/rts_py/rts_input.xlsx")
xls_sheet1 = xls_file1.sheet_by_name("Sheet1")

text_file2 = open("rts_output2.txt", "w")  # uncomment
url_1 = "https://apps.azleg.gov/Account/SignOn" # uncomment

# html_obj = urllib.urlopen(url_1).read()
# soup_obj = BeautifulSoup(html_obj, 'html.parser')
# soup_obj2 = soup_obj.find_all('div')

driver_obj = webdriver.Chrome("chromedriver.exe") # uncomment
# driver_obj2 = webdriver.Opera("D:/apps/WinPython-64bit-3.5.3.1Qt5/bs4_tuts/operadriver_win32/operadriver_win32/operadriver.exe")

# uncomment entire try
try:
    driver_obj.get(url_1)
    # driver_obj2.get(url_1)
    # browser_obj = WebDriverWait(driver_obj, 10).until(EC.presence_of_element_located(element_obj1))
    element_obj1 = driver_obj.find_element_by_name("userName")
    element_obj2 = driver_obj.find_element_by_id("signOnPass")

    element_obj1.send_keys("davidayodele@gmail.com")
    element_obj2.send_keys("******")
    element_obj2.send_keys(Keys.ENTER)

    # Assuming the button has the ID "submit"
    driver_obj.find_element_by_class_name("bsi").click()
    time.sleep(10)
    # browser_obj = WebDriverWait(driver_obj, 10)
    # driver_obj.get(url_1)
    element_obj3 = driver_obj.find_element_by_class_name("billNumberEntry")
    text_file = []
    for i in range(xls_sheet1.nrows):
        text_file.append(open("rts_output" + str(i) + ".txt", "w+"))
        # print xls_sheet1.cell_value(i, 0)
        element_obj3.send_keys(xls_sheet1.cell_value(i, 0))
        element_obj3.send_keys(Keys.ENTER)
        # element_obj3.click()
        # element_obj3.find_element(By.ID("XXXX")).click()
        # element_obj3.submit() throws error in epocrates
        time.sleep(10)
        # driver_obj.find_element_by_tag_name ("#tab1-2").click()
        s1 = driver_obj.page_source
        element_obj3.clear()
        s2 = s1.encode('utf-8').strip()
        text_file[i].write(s2)
    # driver_obj.find_element_by_id("XXXX").click()
    # time.sleep(10)
    # s1 = soup_obj.prettify()
    # time.sleep(10)
finally:
    driver_obj.close()  # end uncomment
#
# print s1

# ======== Parsing ===========
s2_H = s3_H = s4_H = s5_H = s6_H = s7_H = s2_S = s3_S = s4_S = s5_S = s6_S = s7_S = s4_G = s5_G = "NA"
text_files = [None]*xls_sheet1.nrows
for i in range(xls_sheet1.nrows):
    text_files[i] = open("E:/css/crawling/code/rts_py/rts_output" + str(i) + ".txt", "w+")
    soup_obj = soup(text_files[i].read(), "html.parser")  # ,"lxml.parser")  .lower() after read to
    # soup_obj2 = soup_obj.find_all('table')  # creates 2D array of tables

    # ======== House THIRD ==========

    tableH_tags_list = soup_obj.findAll("table", {"class": "table table-bordered table-striped table-bsicondensed"})
    for tableH_tag in tableH_tags_list:
        if (tableH_tag.findAll("button", {"data-target": "#modalFloorCommitteeDetails"}) is not None):
            buttonH_tags_list = tableH_tag.findAll("button", {"data-target": "#modalFloorCommitteeDetails"})
            # print "\n" + "file " + str(i) + "\n"
            # print button_tags_list
            for buttonH_tag in buttonH_tags_list:
                s1_H = buttonH_tag.encode('utf-8').strip()
                expr = regex.compile("Show\s[A-Za-z0-9_\s]+")
                if (expr.search(s1_H) is None) or (expr.search(s1_H).group() != "Show House THIRD"):
                    s2_H = "NA"
                else:
                    s2_H = expr.search(s1_H).group()
                    print "\n" + "file " + str(i) + "\n" + s2_H
                    expr2 = regex.compile("text\s[A-Za-z0-9_\s]+")
                    if(tableH_tag.findAll("label", {"data-bind": lambda val: val and val.startswith("text:")})) is not None:
                        labelH_tags_list = tableH_tag.findAll("label", {"data-bind": "text: (Action=='UA'?Action:Action)"})
                        # print label_tags_list[0].contents
                        labelH_tags_list2 = tableH_tag.findAll("label", {"data-bind": "text: col1"})
                        # print label_tags_list2[0].contents
                        labelH_tags_list3 = tableH_tag.findAll("label", {"data-bind": "text: col2"})
                        # print label_tags_list3[0].contents
                        if labelH_tags_list and labelH_tags_list2 and labelH_tags_list3: # using boolean implicit emptyness
                            if (labelH_tags_list3[0].string is not None) and (labelH_tags_list2[0].string is not None) and (labelH_tags_list[0].string is not None):
                                s5_H = ' '.join(labelH_tags_list3[0].string.split())
                                s4_H = ' '.join(labelH_tags_list2[0].string.split())
                                s3_H = ' '.join(labelH_tags_list[0].string.split())

                        print "Ayes: " + s4_H
                        print "Nays: " + s5_H
                        print s3_H

    labelH_tags_list4 = soup_obj.findAll("label", {"data-bind": lambda val: val and val.startswith("text: ((Body=='S') ? 'Transmit to")})
    labelH_tags_list5 = soup_obj.findAll("label", {"data-bind": "text: moment(SortedDate).format('MM/DD/YYYY')"})

    if labelH_tags_list4 and labelH_tags_list5:  # using boolean implicit emptyness
        if (labelH_tags_list4[0].string is not None) and (labelH_tags_list5[0].string is not None):
            s6_H = ' '.join(labelH_tags_list4[0].string.split())
            s7_H = ' '.join(labelH_tags_list5[0].string.split())
    print s6_H + " " + s7_H

    # ======== Senate FINAL ==========

    tableS_tags_list = soup_obj.findAll("table", {"class": "table table-bordered table-striped table-bsicondensed"})
    for tableS_tag in tableS_tags_list:
        if (tableS_tag.findAll("button", {"data-target": "#modalFloorCommitteeDetails"}) is not None):
            buttonS_tags_list = tableS_tag.findAll("button", {"data-target": "#modalFloorCommitteeDetails"})
            # print "\n" + "file " + str(i) + "\n"
            # print button_tags_list
            for buttonS_tag in buttonS_tags_list:
                s1_S = buttonS_tag.encode('utf-8').strip()
                expr = regex.compile("Show\s[A-Za-z0-9_\s]+")
                if (expr.search(s1_S) is None) or (expr.search(s1_S).group() != "Show Senate FINAL"):
                    s2_S = "NA"
                else:
                    s2_S = expr.search(s1_S).group()
                    print "\n" + "file " + str(i) + "\n" + s2_S
                    expr2 = regex.compile("text\s[A-Za-z0-9_\s]+")
                    if(tableS_tag.findAll("label", {"data-bind": lambda val: val and val.startswith("text:")})) is not None:
                        labelS_tags_list = tableS_tag.findAll("label", {"data-bind": "text: (Action=='UA'?Action:Action)"})
                        # print label_tags_list[0].contents
                        labelS_tags_list2 = tableS_tag.findAll("label", {"data-bind": "text: col1"})
                        # print label_tags_list2[0].contents
                        labelS_tags_list3 = tableS_tag.findAll("label", {"data-bind": "text: col2"})
                        # print label_tags_list3[0].contents
                        if labelS_tags_list and labelS_tags_list2 and labelS_tags_list3:
                            if (labelS_tags_list3[0].string is not None) and (labelS_tags_list2[0].string is not None) and (labelS_tags_list[0].string is not None):
                                s5_S = ' '.join(labelS_tags_list3[0].string.split())
                                s4_S = ' '.join(labelS_tags_list2[0].string.split())
                                s3_S = ' '.join(labelS_tags_list[0].string.split())

                        print "Ayes: " + s4_S
                        print "Nays: " + s5_S
                        print s3_S

    labelS_tags_list4 = soup_obj.findAll("label", {"style": "font-weight:bold"})
    labelS_tags_list5 = soup_obj.findAll("label", {"data-bind": "text: moment(SortedDate).format('MM/DD/YYYY')"})
    if labelS_tags_list4 and labelS_tags_list5:
        if (labelS_tags_list4[0].string is not None) and (labelS_tags_list5[0].string is not None):  # using boolean implicit emptyness
            s6_S = ' '.join(labelS_tags_list4[0].string.split())
            s7_S = ' '.join(labelS_tags_list5[0].string.split())

    print s6_S + " " + s7_S
    # s = s2
    # s4 = s[s.find("<label data-bind=") + 1:s.find("</label>")]
    # print s4

    # ======== Governor FINAL ==========

    divG_tags_list = soup_obj.findAll("div", {"class": "col-md-6"})
    labelG_tags_list = divG_tags_list[1].findAll("label", {"data-bind":"text: ((billInfo() && billInfo().GovernorActionDate != null && billInfo().GovernorActionDate!='') ? moment(billInfo().GovernorActionDate).format('MM/DD/YYYY') : '')"})
    # labelG_tags_list = divG_tag.findAll("label", {"data-bind": "text: ((billInfo() && billInfo().GovernorActionDate != null && billInfo().GovernorActionDate!='') ? moment(billInfo().GovernorActionDate).format('MM/DD/YYYY') : '')"})
    # print labelG_tags_list[0].contents
    labelG_tags_list2 = divG_tags_list[1].findAll("label", {"data-bind": "text: billInfo() ? billInfo().GovernorAction : ''"})
    # print labelG_tags_list2[0].contents
    # labelG_tags_list3 = tableG_tag.findAll("label", {"data-bind": "text: col2"})
    # print labelG_tags_list3[0].contents
    if labelG_tags_list and labelG_tags_list2:
        if (labelG_tags_list[0].string is not None) and (labelG_tags_list2[0].string is not None):
            s5_G = ' '.join(labelG_tags_list[0].string.split())
            s4_G = ' '.join(labelG_tags_list2[0].string.split())

    print s4_G + ": " + s5_G

    # ======== Excel export ==============
    df = pandas.DataFrame([[s2_H, s4_H, s5_H, s3_H, s6_H, s7_H, s2_S, s4_S, s5_S, s3_S, s6_S, s7_S, s4_G, s5_G]])
    #
    # writer = pandas.ExcelWriter('test1.xls')
    # df.to_excel(writer, sheet_name='Sheet1')
    # writer.save()

    # rts_book1 = xlwt.Workbook()
    # rts_sheet1 = rts_book1.add_sheet('sheet1')

    rts_book1 = load_workbook('rts_input.xlsx')
    writer = pandas.ExcelWriter('rts_input.xlsx', engine='openpyxl')
    writer.book = rts_book1
    writer.sheets = dict((ws.title, ws) for ws in rts_book1.worksheets)
    df.to_excel(writer, sheet_name='Sheet1', startrow=i, startcol=2, header=False, index=False)

    # for sheetname, df in df.iteritems():  # loop through `dict` of dataframe cols
    #     df.to_excel(writer, sheet_name=sheetname)  # send df to writer
    #     worksheet = writer.sheets[sheetname]  # pull worksheet object
    #     for idx, col in enumerate(df):  # loop through all columns
    #         series = df[col]
    #         max_len = max((
    #             series.astype(str).map(len).max(),  # len of largest item
    #             len(str(series.name))  # len of column name/header
    #         )) + 1  # adding a little extra space
    #         worksheet.set_column(idx, idx, max_len)  # set column width

    writer.save()
    # for i in range(1, 10, 1):
    #     rts_sheet1.write(i, 2, s2_H)
    #
    # test_sheet1.write(0, 0, "Max Dosage")
    # test_book1_name = "test_book1.xls"
    # test_book1.save(test_book1_name)

    # for table in soup_obj2:
    #     if table.has_attr('class') and table['class'] == ['innerL']:
    #     #     # text_info = div.find_all
    #     #     #print div
    #         tbody_tags_obj = table.find_all('tbody')  # creates 2D array of <p> tag contents (each obj[i] contains ith <p> tag info
    #         print table[0].contents
            # dl_tags_obj = div.find_all('dl')
            # dd_tags_obj = div.find_all('dd')  # creates 2D array of dt tag contents (each obj[i] contains ith <dt> tag info
            # btn btn-sm btn-primary
            # print button.contents
            # print dd_tags_obj[0].contents
            # s2 = s1.replace("\n", " ")
            # for p in p_tags_obj:
            #    print
            # for i in p_tags_obj[0].contents:
            # #    print content
            #     s1 = i.encode('utf-8').strip()  # returns tab: 5mg.. bc Dosage forms is within another tag
            #     # s2 = s1.replace("\n", " ")
            #     print s1
            #     # end for
            # # print dl_tags_obj[0].contents
            # for i in dd_tags_obj[0].contents:
            #     s2 = i.encode('utf-8').strip()
            #     print s2[0:19]
            #     # reg = re.compile("[a-z]+8?")
            #     # str = "ccc8"
            #     # print(reg.match(str).group())
            #     expr = regex.compile("max: \d\d mg/day")
            #     s3 = expr.search(s2).group()
            #     print s3
    text_files[i].close()

# xls_file1.close() # closed automatically by constructor
# with open("textFileWithHtml.txt") as markup:
#     soup = BeautifulSoup(markup.read())
#
# with open("strip_textFileWithHtml.txt", "w") as f:
#     f.write(soup.get_text().encode('utf-8'))
